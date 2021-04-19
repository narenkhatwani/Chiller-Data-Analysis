from tkinter import *
from tkinter import messagebox
import re,pymysql
from tkinter import Tk
from PIL import ImageTk, Image
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.dates as mdates
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import LabelEncoder
from openpyxl import *


data_naren=pd.read_csv("C:\\Datasets\\internshipdataset1.csv")
data_harsh=pd.read_csv("C:\\Datasets\\internshipdataset1.csv")
data_priya=pd.read_csv("C:\\Datasets\\Dataset2.csv")#reading csv file



#function1
plt.subplots(figsize=(6,6))
plt.title("THE PERCENTAGE OF DOWNLOADS CATEGORY WISE")
sns.countplot(data_naren['Category'],label='percentage')
plt.xticks(fontsize=6,rotation=78)
plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function1.jpg',dpi=100)
#function1

#function5
data_naren['Year'] = pd.DatetimeIndex(data_naren['Last Updated']).year
plt.subplots(figsize=(6,6))
plt.title("DOWNLOAD TREND CATEGORY WISE")
sns.countplot(x=data_naren['Category'],hue=data_naren['Year'])
plt.xticks(fontsize=6,rotation=78)
plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function5.jpg',dpi=100)
#function5

#function6
data_naren1=data_naren[(data_naren['Year']>=2015)&(data_naren['Year']<=2018)]
plt.subplots(figsize=(6,6))
plt.title("DOWNLOAD TREND CATEGORY WISE")
sns.countplot(x="Category",hue='Year',data=data_naren1)
plt.xticks(fontsize=6,rotation=78)
plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function6.jpg',dpi=100)
#function6

#function7
plt.subplots(figsize=(6,6))
plt.title("THE PERCENTAGE OF DOWNLOADS AS ANDROID VERSION")
sns.countplot(data_naren['Android Ver'],label='percentage')
plt.xticks(fontsize=6,rotation=78)
plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function7.jpg',dpi=100)
#function7

#function8
data_naren['Installs']=data_naren['Installs'].map(lambda x:x.rstrip('+'))
data_naren['Installs']=data_naren['Installs'].map(lambda x:''.join(x.split(',')))

#function8
lb_make=LabelEncoder()

data_naren['Category NUM']=lb_make.fit_transform(data_naren['Category'])
print(data_naren['Category NUM'])
X=data_naren[["Category NUM"]]
#c=data1[data1['Category NUM'][1],data1['Category NUM'][4],data1['Category NUM'][20],data1['Category NUM'][8],data1['Category NUM'][16],data1['Category NUM'][24],data1['Category NUM'][30],data1['Category NUM'][28],data1['Category NUM'][12],data1['Category NUM'][18]]
#X=data1[c]
y=data_naren['Installs']

X_train,X_test,y_train,y_test=train_test_split(X,y,test_size=0.30,random_state=0)


regressor=LinearRegression()
regressor.fit(X_train,y_train)
y_pred=regressor.predict(X_test)

data1=pd.DataFrame({'Actual':y_test,'Predicted':y_pred})
print(data1)

plt.title("PREDICTION")
plt.scatter(y_test,y_pred)
plt.xlabel("DOWNLOADS")
plt.ylabel("PREDICTED DOWNLOADS")
plt.xticks(rotation=90)
plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function8.jpg',dpi=100)
#function8

#function10
data_naren['Month'] = pd.DatetimeIndex(data_naren['Last Updated']).month
plt.subplots(figsize=(5,5))
plt.title("THE DOWNLOADS AS PER MONTH")
sns.countplot(x="Month",data=data_naren)
plt.xticks(fontsize=8,rotation=78)
plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function10.jpg',dpi=100)    
#function10

#function11
data_naren['quarter_year'] = pd.to_datetime(data_naren['Last Updated']).dt.to_period('Q')
plt.subplots(figsize=(6,6))
plt.title("THE DOWNLOADS AS PER QUARTER")
sns.countplot(data_naren['Category'],hue=data_naren['quarter_year'])
plt.xticks(fontsize=8,rotation=85)
plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function11.jpg',dpi=100)
#function11
#function19 extra
plt.title("THE DOWNLOADS AS PER QUARTER CATEGORY")
plt.subplots(figsize=(6,6))
sns.countplot(data_naren['Category'],hue=data_naren['quarter_year'])
plt.xticks(fontsize=8,rotation=90)
plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function19.jpg',dpi=100)
#function19 extra
# This function is used for adjusting window size and making the necessary configuration on start of window 
def adjustWindow(window): 
    w = 800  # width for the window size 
    h = 800  # height for the window size 
    ws = screen.winfo_screenwidth()  # width of the screen 
    hs = screen.winfo_screenheight()  # height of the screen 
    x = (ws/2) - (w/2)  # calculate x and y coordinates for the Tk window 
    y = (hs/2) - (h/2) 
    window.geometry('%dx%d+%d+%d' % (w, h, x, y))  # set the dimensions of the screen and where it is placed 
    window.resizable(False, False)    # disabling the resize option for the window 
    window.configure(background='white')    # making the background white of the window


# This function is used for adjusting window size and making the necessary configuration on start of window 
def adjustWindow1(window): 
    w = 800  # width for the window size 
    h = 530
    ws = screen.winfo_screenwidth()  # width of the screen 
    hs = screen.winfo_screenheight()  # height of the screen 
    x = (ws/2) - (ws/2)  # calculate x and y coordinates for the Tk window 
    y = (hs/2) - (hs/2) 
    window.geometry('%dx%d+%d+%d' % (w, h, x, y))  # set the dimensions of the screen and where it is placed 
    window.resizable(False, False)    # disabling the resize option for the window 
    window.configure(background='white')    # making the background white of the window

# This function is used for adjusting window size and making the necessary configuration on start of window 
def adjustWindow2(window): 
    w = 800  # width for the window size 
    h = 600
    ws = screen.winfo_screenwidth()  # width of the screen 
    hs = screen.winfo_screenheight()  # height of the screen 
    x = (ws/2) - (ws/2)  # calculate x and y coordinates for the Tk window 
    y = (hs/2) - (hs/2) 
    window.geometry('%dx%d+%d+%d' % (w, h, x, y))  # set the dimensions of the screen and where it is placed 
    window.resizable(True, True)    # disabling the resize option for the window 
    window.configure(background='white')    # making the background white of the window


# registration window 
def register(): 
    global screen1, fullname, email, password, repassword, country, gender, tnc # making all entry field variable global 
    fullname = StringVar() 
    email = StringVar() 
    password = StringVar() 
    repassword = StringVar() 
    gender = StringVar() 
    country = IntVar() 
    tnc = IntVar() 
    screen1 = Toplevel(screen) 
    screen1.title("BRUTE FORCE") 
    adjustWindow(screen1) # configuring the window 
    Label(screen1, text="Registration Form", width='55', height="2", font=("Calibri", 22, 'bold'), fg='#d9660a').place(x=0, y=0)
    #Label(screen1, text="", bg='#174873', width='55', height='30').place(x=105, y=100) 
    
    photo = PhotoImage(file="reg.png") # opening left side image - Note: If image is in same folder then no need to mention the full path 
    label = Label(screen1, image=photo, text="") # attaching image to the label 
    label.place(x=-1, y=80) 
    label.image = photo # it is necessary in Tkinter to keep a instance of image to display image in label 

   
    #Label(screen1, text="REGISTRATION FORM", font=("Open Sans", 22, 'bold'), fg='white', bg='#174873', anchor=W).place(x=250, y=30) 
    #Label(screen1, text="Registration Form", width='55', height="1", font=("Calibri", 22, 'bold'), fg='#d9660a').place(x=0, y=20) 
    Label(screen1, text="Full Name:", font=("Open Sans", 11, 'bold'), fg='white', bg='#174873', anchor=W).place(x=350, y=130) 
    Entry(screen1, textvar=fullname).place(x=480, y=130) 
    Label(screen1, text="Email ID:", font=("Open Sans", 11, 'bold'), fg='white', bg='#174873', anchor=W).place(x=350, y=180) 
    Entry(screen1, textvar=email).place(x=480, y=180) 
    Label(screen1, text="Country:", font=("Open Sans", 11, 'bold'), fg='white', bg='#174873', anchor=W).place(x=350, y=230) 
    Radiobutton(screen1, text="India", variable=country, value=1,fg='white', bg='#174873').place(x=480, y=230)     
    Radiobutton(screen1, text="Other", variable=country, value=2,fg='white', bg='#174873').place(x=550, y=230) 
                    
    Label(screen1, text="gender:", font=("Open Sans", 11, 'bold'), fg='white', bg='#174873', anchor=W).place(x=350, y=280) 
    list1 = ['Male','Female','Other']
    droplist = OptionMenu(screen1, gender, *list1) 
    droplist.config(width=18) 
    gender.set('--select your gender--') 
    droplist.place(x=480, y=275) 
    Label(screen1, text="Password:", font=("Open Sans", 11, 'bold'), fg='white',bg='#174873', anchor=W).place(x=350, y=330) 
    Entry(screen1, textvar=password, show="*").place(x=480, y=330) 
    Label(screen1, text="Re-Password:", font=("Open Sans", 11, 'bold'), fg='white',bg='#174873', anchor=W).place(x=350, y=380) 
    entry_4 = Entry(screen1, textvar=repassword, show="*") 
    entry_4.place(x=480, y=380) 
    Checkbutton(screen1, text="I accept all terms and conditions", variable=tnc, font=("Open Sans", 9, 'bold'), fg='brown').place(x=375, y=430) 
    Button(screen1, text='Submit', width=15, font=("Open Sans", 13, 'bold'), bg='brown', fg='black',command=register_user).place(x=330, y=480)              
    Button(screen1, text='Back', width=10, font=("Open Sans", 13, 'bold'), bg='brown', fg='black',command=screen1.destroy).place(x=550, y=480)              

    photo = PhotoImage(file="log.png") # opening left side image - Note: If image is in same folder then no need to mention the full path 
    label = Label(screen1, image=photo, text="") # attaching image to the label 
    label.place(x=130, y=0) 
    label.image = photo # it is necessary in Tkinter to keep a instance of image to display image in label 
    
         

def register_user(): 
    if fullname.get() and email.get() and password.get() and repassword.get() and country.get(): # checking for all empty values in entry field 
        if gender.get() == "--select your gender--": # checking for selection of university 
            Label(screen1, text="Please select your gender", fg="red", 
                  font=("calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) 
            return 
        else: 
            if tnc.get(): # checking for acceptance of agreement 
                if re.match("^.+@(\[?)[a-zA-Z0-9-.]+.([a-zA-Z]{2,3}|[0-9]{1,3})(]?)$", email.get()): # validating the email 
                    if password.get() == repassword.get(): # checking both password match or not 
                        # if u enter in this block everything is fine just enter the values in database 
                        country_value = 'India' 
                        if country.get() == 2:  
                            
                            country_value = 'Other' 
                        connection = pymysql.connect(host="localhost", user="root", passwd="", database="bruteforce") # database connection 
                        cursor = connection.cursor() 
                        insert_query = "INSERT INTO registration_details (fullname, email, password, country,gender) VALUES('"+ fullname.get() + "', '"+ email.get() + "', '"+ password.get() + "', '"+ country_value + "', '"+ gender.get() + "' );" # queries for inserting values 
                        cursor.execute(insert_query) # executing the queries 
                        connection.commit() # commiting the connection then closing it. 
                        connection.close() # closing the connection of the database 
                        Label(screen1, text="Registration Sucess", fg="green", font=( "calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) # printing successful registration message 
                        Button(screen1, text='Proceed to Login ->', width=20, font=("Open Sans", 9, 'bold'), bg='brown', fg='white',command=screen1.destroy).place(x=170, y=565) # button to navigate back to login page 
                    else: 
                        Label(screen1, text="Password does not match", fg="red", font=( "calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) 
                        return 
                else: 
                    Label(screen1, text="Please enter valid email id", fg="red", font=("calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) 
                    return 
            else: 
                Label(screen1, text="Please accept the agreement", fg="red", 
                      font=("calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) 
                return 
    else: 
        Label(screen1, text="Please fill all the details", fg="red", 
        font=("calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) 
        return 

# login creditentials verification 
def login_verify(): 
    global registrationID 
    connection = pymysql.connect(host="localhost", user="root", passwd="", database="bruteforce") # database connection 
    cursor = connection.cursor() 
    select_query =  "SELECT * FROM registration_details where email = '" + username_verify.get() + "' AND password = '" + password_verify.get() + "';" # queries for retrieving values 
    cursor.execute(select_query) # executing the queries 
    registration_info = cursor.fetchall() 
    connection.commit() # commiting the connection then closing it. 
    connection.close() # closing the connection of the database                     
    if registration_info: 
        messagebox.showinfo("Congratulation", "Login Succesfull") # displaying message for successful login 
        registrationID = registration_info[0][0] 
        welcome_page(registration_info) # opening welcome window 
    else: 
        messagebox.showerror("Error", "Invalid Username or Password") # displaying message for invalid details 


# welcome window 
def welcome_page(registration_info): 
    global screen2 
    screen2 = Toplevel(screen) 
    screen2.title("BRUTE FORCE") 
    adjustWindow1(screen2) # configuring the window 
    Label(screen2, text="Welcome " + registration_info[0][1], width='47', height="2", font=("Calibri", 25, 'bold'), fg='white', bg='#d9660a').place(x=0, y=0) 
    Label(screen2, text="", bg='#174873', width='20', height='20').place(x=0, y=96) 
    Message(screen2, text='“If we have data, let’s look at data. If all we have are opinions, let’s go with mine.”\n\n - — Jim Barksdale', width='100', font=("Helvetica", 10, 'bold', 'italic'), fg='white', bg='#174873', anchor = CENTER).place(x=10, y=100) 
    
    #photo = PhotoImage(file="bft.png") # opening left side image - Note: If image is in same folder then no need to mention the full path 
    #label = Label(screen2, image=photo, text="") # attaching image to the label 
    #label.place(x=10, y=270) 
    #label.image = photo # it is necessary in Tkinter to keep a instance of image to display image in label 
    
    photo1 = PhotoImage(file="analy.png") # opening right side image - Note: If image is in same folder then no need to mention the full path 
    label1 = Label(screen2, image=photo1, text="") # attaching image to the label 
    label1.place(x=150, y=96) 
    label1.image = photo1 # it is necessary in Tkinter to keep a instance of image to display image in label 
    
    photo1 = PhotoImage(file="pay.png") # opening right side image - Note: If image is in same folder then no need to mention the full path 
    label1 = Label(screen2, image=photo1, text="") # attaching image to the label 
    label1.place(x=180, y=0) 
    label1.image = photo1 # it is necessary in Tkinter to keep a instance of image to display image in label 
  
    
    Button(screen2, text='Fetch Record', width=20,height=4, font=("Open Sans", 13, 'bold'), bg='brown', fg='white',command=function14).place(x=270, y=150) 
    Button(screen2, text='Insert Record', width=20,height=4, font=("Open Sans", 13, 'bold'), bg='brown', fg='white',command=add_data).place(x=270, y=250) 
    Button(screen2, text='Perform Function', width=20,height=4, font=("Open Sans", 13, 'bold'), bg='brown', fg='white',command=next_page).place(x=270, y=350)
    Button(screen2, text='Back', width=10, font=("Open Sans", 13, 'bold'), bg='brown', fg='black',command=screen2.destroy).place(x=380, y=500)              

    
def next_page():    
    global screen3 
    screen3 = Toplevel(screen) 
    screen3.title("BRUTE FORCE")
    adjustWindow2(screen3)
    #screen3.geometry('600*600')
    photo1 = PhotoImage(file="bigdata.png") # opening right side image - Note: If image is in same folder then no need to mention the full path 
    label1 = Label(screen3, image=photo1, text="") # attaching image to the label 
    label1.place(x=0, y=0) 
    label1.image = photo1 # it is necessary in Tkinter to keep a instance of image to display image in label 
    
    Label(screen3, text="""
    1) What is the percentage download in each category on the playstore.
    2) How many apps have managed to get the following number of downloads
    a) Between 10,000 and 50,000 b) Between 50,000 and 150000 c) Between 150000 and 500000
    d) Between 500000 and 5000000 e) More than 5000000
    3) Which category of apps have managed to get the most,least and an average of 2,50,000 downloads atleast.
    4) Which category of apps have managed to get the highest maximum average ratings from the users.
    5) What is the download trend category wise over the period for which the data is being made available.
    6) For the years 2016,2017,2018 what are the category of apps that have got the most and the least downloads. What is the percentage
    increase or decrease that the apps have got over the period of three years
    7) All those apps , whose android version is not an issue and can work with varying devices ,what is the percentage
    increase or decrease in the downloads.
    8) Amongst sports, entertainment,social media,news,events,travel and games,which is the category of app that is most
    likely to be downloaded in the coming years, kindly make a prediction and back it with suitable findings.
    9) All those apps who habve managed to get over 1,00,000 downloads, have they managed to get an average rating
    of 4.1 and above? An we conclude something in co-relation to the number of downloads and the ratings received.
    10) Across all the years ,which month has seen the maximum downloads fr each of the category. What is the ratio
    of downloads for the app that qualifies as teen versus mature17+
    11) Which quarter of which year has generated the highest number of install for each app used in the study?
    12) Which of all the apps given have managed to generate the most positive and negative sentiments.Also figure 
    out the app which has generated approximately the same ratio for positive and negative sentiments.
    13) Study and find out the relation between the Sentiment-polarity and sentiment-subjectivity of all the apps.
    14) Generate an interface where the client can see the reviews categorized as positive.negative and neutral 
    ,once they have selected the app from a list of apps available for the study.
    15) Is it advisable to launch an app like ’10 Best foods for you’? Do the users like these apps?
    16) Which month(s) of the year , is the best indicator to the avarage downloads that an app will generate over the entire year?
    17) Does the size of the App influence the number of installs that it gets ? if,yes the trend is positive or negative with the 
    increase in the app size.
    18) Provide an interface to add new data to both the datasets provided.
    
                  """,fg='black').place(x=10,y=10)
    Button(screen3,text='Function 1',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function1).place(x=720,y=20)
    Button(screen3,text='Function 2',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function2).place(x=720,y=50)
    Button(screen3,text='Function 3',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function3).place(x=720,y=80)
    Button(screen3,text='Function 4',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function4).place(x=720,y=110)
    Button(screen3,text='Function 5',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function5).place(x=720,y=140)
    Button(screen3,text='Function 6',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function6).place(x=720,y=170)
    Button(screen3,text='Function 7',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function7).place(x=720,y=200)
    Button(screen3,text='Function 8',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function8).place(x=720,y=230)
    Button(screen3,text='Function 9',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function9).place(x=720,y=260)
    Button(screen3,text='Function 10',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function10).place(x=720,y=290)
    Button(screen3,text='Function 11',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function11).place(x=720,y=320)
    Button(screen3,text='Function 12',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function12).place(x=720,y=350)
    Button(screen3,text='Function 13',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function13).place(x=720,y=380)
    Button(screen3,text='Function 15',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function15).place(x=720,y=410)
    Button(screen3,text='Function 16',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function16).place(x=720,y=440)
    Button(screen3,text='Function 17',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function17).place(x=720,y=470)
    Button(screen3,text='Function 4b',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function4b).place(x=620,y=110)
    Button(screen3,text='Function 19',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function19).place(x=620,y=440)
    Button(screen3,text='Function 20',width=9,height=1,font=("Open Sans",9,'bold'),bg='brown',fg='white',command=function20).place(x=620,y=470)
    

def main_screen():
    global screen, username_verify, password_verify
    screen = Tk()
    username_verify = StringVar()
    password_verify = StringVar()
    screen.title("BRUTE FORCE")
    adjustWindow(screen)
    
    img2 = ImageTk.PhotoImage(Image.open("bigdata.png"))  
    d1 = Label(image= img2) 
    d1.place(x =-1,y = 70)
    
    Label(screen,text="BRUTE FORCE Manager",width="500",height="2",font=("Calibri",22,'bold'),fg='white',bg='#d9660a').pack()
    Label(screen, text="Please enter details below to login", bg='#174873', fg='white').place(x=312,y=150) 
    Label(screen, text="Username * ", font=("Open Sans", 10, 'bold'), bg='#174873', fg='white').place(x=358,y=192)             
    Entry(screen, textvar=username_verify).place(x=337,y=215) 
    Label(screen, text="Password * ", font=("Open Sans", 10, 'bold'), bg='#174873', fg='white').place(x=358,y=262) 
    Entry(screen, textvar=password_verify, show="*").place(x=337,y=285) 
    Button(screen, text="LOGIN", bg="#e79700", width=15, height=1, font=("Open Sans", 13, 'bold'), fg='white', command=login_verify).place(x=320,y=325) 
    Button(screen, text="New User? Register Here", height="2", width="30", bg='#e79700', font=("Open Sans", 10, 'bold'), fg='white', command=register).place(x=270,y=380) 
    
    img1 = ImageTk.PhotoImage(Image.open("bft.png"))  
    c1 = Label(image= img1) 
    c1.place(x =50,y = 0) 
    
             
    screen.mainloop()
#
def add_data():
    global appname,category,rating,review,size,installs,types,price,content_rating,genres,last_update,current_ver,android_ver,tnc
    root=Tk()
    root.geometry('1000x1000')
    root.title("Registration form")
    appname=StringVar()
    category=StringVar()
    rating=StringVar()
    review=StringVar()
    size=StringVar()
    installs=StringVar()
    types=IntVar()
    price=StringVar()
    content_rating=StringVar()
    genres=StringVar()
    last_update=StringVar()
    current_ver=StringVar()
    android_ver=StringVar()
    tnc=IntVar()    
    
    label_0=Label(root,text="Modify Dataset 1",width=20,font=("bold",20))
    label_0.place(x=90,y=53)

    label_1=Label(root,text="App Name",width=20,font=("bold",10))
    label_1.place(x=80,y=130)

    entry_1=Entry(root,textvar=appname)
    entry_1.place(x=240,y=130)
    
    
    label_4=Label(root,text="Category",width=20,font=("bold",10))
    label_4.place(x=70,y=180)
    
    list1=['ART_AND_DESIGN','AUTO_AND_VEHICLES','BEAUTY','BOOKS_AND_REFERENCE','BUSINESS','COMICS','COMMUNICATION','DATING','EDUCATION','ENTERTAINMENT','EVENTS','FOOD_AND_DRINK','HEALTH_AND_FITNESS','HOUSE_AND_HOME','LIBRARIES_AND_DEMO','LIFESTYLE','GAME','FAMILY','MEDICAL','SHOPPING','PHOTOGRAPHY','SPORTS','TRAVEL_AND_LOCAL','TOOLS','PERSONALIZATION','PRODUCTIVITY','PARENTING','WEATHER','VIDEO_PLAYERS','NEWS_AND_MAGAZINES','MAPS_AND_NAVIGATION','SOCIAL','FINANCE']
    c=StringVar()
    droplist=OptionMenu(root,c,*list1)
    droplist.config(width=15)
    c.set('Select your category')
    droplist.place(x=240,y=180)
    

    label_2=Label(root,text="Rating",width=20,font=("bold",10))
    label_2.place(x=68,y=230)
    
    entry_2=Entry(root,textvar=rating)
    entry_2.place(x=240,y=230)

    label_3=Label(root,text="Review",width=20,font=("bold",10))
    label_3.place(x=68,y=280)

    entry_3=Entry(root,textvar=review)
    entry_3.place(x=240,y=280)

    label_4=Label(root,text="Size",width=20,font=("bold",10))
    label_4.place(x=68,y=330)
    
    entry_4=Entry(root,textvar=size)
    entry_4.place(x=240,y=330)

    label_5=Label(root,text="Installs",width=20,font=("bold",10))
    label_5.place(x=68,y=380)

    entry_5=Entry(root,textvar=installs)
    entry_5.place(x=240,y=380)
    


    label_6=Label(root,text="Type",width=20,font=("bold",10))
    label_6.place(x=70,y=430)
    
    Radiobutton(root,text="Free",padx=5,variable=types,value=1).place(x=235,y=430)
    Radiobutton(root,text="Paid",padx=20,variable=types,value=2).place(x=290,y=430)

    label_7=Label(root,text="Price",width=20,font=("bold",10))
    label_7.place(x=68,y=470) 

    entry_2=Entry(root,textvar=price)
    entry_2.place(x=240,y=470)


    label_8=Label(root,text="Content Rating",width=20,font=("bold",10))
    label_8.place(x=70,y=520)

    list2=['Everyone','Everyone 10+','Teen','Mature 17+']
    d=StringVar()
    droplist=OptionMenu(root,d,*list2)
    droplist.config(width=15)
    d.set('Select your category')
    droplist.place(x=240,y=520)

    label_9=Label(root,text="Genres",width=20,font=("bold",10))
    label_9.place(x=68,y=570)

    entry_9=Entry(root,textvar=genres)
    entry_9.place(x=240,y=570)
    

    label_10=Label(root,text="Last Updated",width=20,font=("bold",10))
    label_10.place(x=68,y=620)

    entry_10=Entry(root,textvar=last_update)
    entry_10.place(x=240,y=620)


    label_11=Label(root,text="Current Version",width=20,font=("bold",10))
    label_11.place(x=68,y=650)

    entry_11=Entry(root,textvar=current_ver)
    entry_11.place(x=240,y=650)


    label_12=Label(root,text="Android Version",width=20,font=("bold",10))
    label_12.place(x=68,y=700)

    entry_12=Entry(root,textvar=android_ver)
    entry_12.place(x=240,y=700)
    
    
    Checkbutton(root, text="I accept all terms and conditions", variable=tnc, bg='#174873', font=("Open Sans", 9, 'bold'), fg='brown').place(x=475, y=430) 
    Button(root, text='Submit', width=20, font=("Open Sans", 13, 'bold'), bg='brown', fg='black',command=insert).place(x=470, y=480)              
       
       
    
    
def validate_submit():
    if appname.get() and category.get() and rating.get() and review.get() and size.get() and installs.get() and types.get() and price.get() and content_rating.get() and genres.get() and last_update.get() and current_ver.get() and android_ver.get() and tnc.get(): 
        if category.get() == "--select your category--": # checking for selection of university 
            Label(root, text="Please select category", fg="red", 
                  font=("calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=730) 
            return
                
        else: 
            
            if tnc.get(): # checking for acceptance of agreement
                if appname.get==appname.isalpha():
                     #add code for adding data to csv file
                        '''
                         current_row = sheet.max_row
                            current_column = sheet.max_column
                            sheet.cell(row=current_row+1,column=5).value = name_field.get()
                            sheet.cell(row=current_row+1,column=7).value = age_field.get()
                            sheet.cell(row=current_row+1,column=6).value = gender_field.get()
                            sheet.cell(row=current_row+1,column=12).value = height_field.get()
                            sheet.cell(row=current_row+1,column=11).value = weight_field.get()
                            sheet.cell(row=current_row,column=8).value = bloodgrp_field.get()
                            sheet.cell(row=current_row,column=15).value = bpsystolic_field.get()
                            sheet.cell(row=current_row,column=16).value = bpdystolic_field.get()
                            sheet.cell(row=current_row,column=10).value = haemoglobin_field.get()
                            sheet.cell(row=current_row,column=13).value = insulinBF_field.get()
                            sheet.cell(row=current_row,column=14).value = insulinAF_field.get()
                            sheet.cell(row=current_row,column=2).value = diabetic_field.get()
                            sheet.cell(row=current_row,column=1).value = stage_field.get()
                            sheet.cell(row=current_row+1, column=9).value = cost_field.get()

                        country_value = 'India' 
                        if country.get() == 2:  
                             
                            country_value = 'Other' 
                        connection = pymysql.connect(host="localhost", user="root", passwd="", database="bruteforce") # database connection 
                        cursor = connection.cursor() 
                        insert_query = "INSERT INTO registration_details (fullname, email, password, country,gender) VALUES('"+ fullname.get() + "', '"+ email.get() + "', '"+ password.get() + "', '"+ country_value + "', '"+ gender.get() + "' );" # queries for inserting values 
                        cursor.execute(insert_query) # executing the queries 
                        connection.commit() # commiting the connection then closing it. 
                        connection.close() # closing the connection of the database 
                        Label(screen1, text="Registration Sucess", fg="green", font=( "calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) # printing successful registration message 
                        Button(screen1, text='Proceed to Login ->', width=20, font=("Open Sans", 9, 'bold'), bg='brown', fg='white',command=screen1.destroy).place(x=170, y=565) # button to navigate back to login page''' 
                
            else: 
                Label(screen1, text="Please accept the agreement", fg="red", 
                      font=("calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) 
                return 
    else: 
        Label(screen1, text="Please fill all the details", fg="red", 
        font=("calibri", 11), width='30', anchor=W, bg='white').place(x=0, y=570) 
        return
#    
def function1():
    global screen4 
    screen4 = Toplevel(screen) 
    screen4.title("FUNCTION 1")
    adjustWindow(screen4)
      
    img = ImageTk.PhotoImage(Image.open('function1.jpg'))
    panel =Label(screen4, image = img)
    panel.pack(side = "bottom", fill = "both", expand = "yes")
        
    screen4.mainloop()

def function2():
    global screen5 
    screen5 = Toplevel(screen) 
    screen5.title("FUNCTION 5")
    adjustWindow(screen5)
    
    data_naren['Installs']=data_naren['Installs'].map(lambda x:x.rstrip('+'))
    data_naren['Installs']=data_naren['Installs'].map(lambda x:''.join(x.split(',')))
    
    #10,000 to 50,000
    x1=data_naren.loc[(data_naren['Installs']=='10000'),['Installs','App']].count() 
    y1=data_naren.loc[(data_naren['Installs']=='50000'),['Installs','App']].count() 
    
    Label(screen5, text="The app downloads 10,000 to 50,000:", bg='yellow', fg='black').pack()
    Label(screen5, text=x1+y1, bg='yellow', fg='black').pack()
    
    
        
    #50,000 to 150,000
    x2=data_naren.loc[(data_naren['Installs']=='50000'),['Installs','App']].count() 
    y2=data_naren.loc[(data_naren['Installs']=='150000'),['Installs','App']].count() 
    
    Label(screen5, text="The app downloads 50,000 to 150,000:", bg='yellow', fg='black').pack()
    Label(screen5, text=x2+y2, bg='yellow', fg='black').pack()
    
    #50,000 to 150,000 ends here
    
    
    #150000 to 500000
    x3=data_naren.loc[(data_naren['Installs']=='150000'),['Installs','App']].count() 
    y3=data_naren.loc[(data_naren['Installs']=='500000'),['Installs','App']].count() 
    
    Label(screen5, text="The app downloads 150000 to 500000:", bg='yellow', fg='black').pack()
    Label(screen5, text=x3+y3, bg='yellow', fg='black').pack()
    
    
     
    y4=data_naren.loc[(data_naren['Installs']=='5000000'),['Installs','App']].count() 
    
    Label(screen5, text="The app downloads above 50 lakhs is:", bg='yellow', fg='black').pack()
    Label(screen5, text=y4, bg='yellow', fg='black').pack()
    

    screen5.mainloop()
    
def function3():
    global screen6 
    screen6 = Toplevel(screen) 
    screen6.title("FUNCTION 3")
    adjustWindow(screen6)
    
    x1=data_naren.groupby(['Category']).agg({'Installs': [min, max]})

    Label(screen6, text=x1, bg='yellow', fg='black').pack()
    
    scren6.mainloop()
    
    
def function4():
    global screen7 
    screen7 = Toplevel(screen) 
    screen7.title("FUNCTION 4")
    adjustWindow(screen7)
      
    x1=data_naren.loc[(data_naren['Category']=='ART_AND_DESIGN')&(data_naren['Rating']),['App','Rating']].mean()
    x2=data_naren.loc[(data_naren['Category']=='AUTO_AND_VEHICLES')&(data_naren['Rating']),['App','Rating']].mean()
    x3=data_naren.loc[(data_naren['Category']=='BEAUTY')&(data_naren['Rating']),['App','Rating']].mean()
    x4=data_naren.loc[(data_naren['Category']=='BOOKS_AND_REFERENCE')&(data_naren['Rating']),['App','Rating']].mean()
    x5=data_naren.loc[(data_naren['Category']=='BUSINESS')&(data_naren['Rating']),['App','Rating']].mean()
    x6=data_naren.loc[(data_naren['Category']=='COMICS')&(data_naren['Rating']),['App','Rating']].mean()
    x7=data_naren.loc[(data_naren['Category']=='COMMUNCIATION')&(data_naren['Rating']),['App','Rating']].mean()
    x8=data_naren.loc[(data_naren['Category']=='DATING')&(data_naren['Rating']),['App','Rating']].mean()
    x9=data_naren.loc[(data_naren['Category']=='EDUCATION')&(data_naren['Rating']),['App','Rating']].mean()
    x10=data_naren.loc[(data_naren['Category']=='ENTERTAINMENT')&(data_naren['Rating']),['App','Rating']].mean()
    x11=data_naren.loc[(data_naren['Category']=='EVENTS')&(data_naren['Rating']),['App','Rating']].mean()
    x12=data_naren.loc[(data_naren['Category']=='FAMILY')&(data_naren['Rating']),['App','Rating']].mean()
    x13=data_naren.loc[(data_naren['Category']=='FINANCE')&(data_naren['Rating']),['App','Rating']].mean()
    x14=data_naren.loc[(data_naren['Category']=='FOOD_AND_DRINK')&(data_naren['Rating']),['App','Rating']].mean()
    x15=data_naren.loc[(data_naren['Category']=='GAME')&(data_naren['Rating']),['App','Rating']].mean()
    x16=data_naren.loc[(data_naren['Category']=='HEALTH_AND_FITNESS')&(data_naren['Rating']),['App','Rating']].mean()
    x17=data_naren.loc[(data_naren['Category']=='HOUSE_AND_HOME')&(data_naren['Rating']),['App','Rating']].mean()
    x18=data_naren.loc[(data_naren['Category']=='LIBRARIES)AND_DEMO')&(data_naren['Rating']),['App','Rating']].mean()
    x19=data_naren.loc[(data_naren['Category']=='LIFESTYLE')&(data_naren['Rating']),['App','Rating']].mean()
       
    
    Label(screen7, text=x1, bg='yellow', fg='black').place()
    Label(screen7, text=x2, bg='yellow', fg='black').pack()
    Label(screen7, text=x3, bg='yellow', fg='black').pack()
    Label(screen7, text=x4, bg='yellow', fg='black').pack()
    Label(screen7, text=x5, bg='yellow', fg='black').pack()
    
    Label(screen7, text=x6, bg='yellow', fg='black').pack()
    Label(screen7, text=x7, bg='yellow', fg='black').pack()
    Label(screen7, text=x8, bg='yellow', fg='black').pack()
    Label(screen7, text=x9, bg='yellow', fg='black').pack()
    Label(screen7, text=x10, bg='yellow', fg='black').pack()
    Label(screen7, text=x11, bg='yellow', fg='black').pack()
    Label(screen7, text=x12, bg='yellow', fg='black').pack()
    Label(screen7, text=x13, bg='yellow', fg='black').pack()
    Label(screen7, text=x14, bg='yellow', fg='black').pack()
    Label(screen7, text=x15, bg='yellow', fg='black').pack()
    Label(screen7, text=x16, bg='yellow', fg='black').pack()
    Label(screen7, text=x17, bg='yellow', fg='black').pack()
    Label(screen7, text=x18, bg='yellow', fg='black').pack()
    Label(screen7, text=x19, bg='yellow', fg='black').pack()
    
    
    
    screen7.mainloop()
def function4b():
    screen28 = Toplevel(screen) 
    screen28.title("FUNCTION 4b")
    adjustWindow(screen28)
    
    x20=data_naren.loc[(data_naren['Category']=='MAPS_AND_NAVIGATION')&(data_naren['Rating']),['App','Rating']].mean()
    x21=data_naren.loc[(data_naren['Category']=='MEDICAL')&(data_naren['Rating']),['App','Rating']].mean()
    x22=data_naren.loc[(data_naren['Category']=='NEWS_AND_MAGAZINE')&(data_naren['Rating']),['App','Rating']].mean()
    x23=data_naren.loc[(data_naren['Category']=='PARENTING')&(data_naren['Rating']),['App','Rating']].mean()
    x24=data_naren.loc[(data_naren['Category']=='PERSONALIZATION')&(data_naren['Rating']),['App','Rating']].mean()
    x24=data_naren.loc[(data_naren['Category']=='PHOTOGRAPHY')&(data_naren['Rating']),['App','Rating']].mean()
    x26=data_naren.loc[(data_naren['Category']=='PRODUCTIVITY')&(data_naren['Rating']),['App','Rating']].mean()
    x27=data_naren.loc[(data_naren['Category']=='SHOPPING')&(data_naren['Rating']),['App','Rating']].mean()
    x28=data_naren.loc[(data_naren['Category']=='SOCIAL')&(data_naren['Rating']),['App','Rating']].mean()
    x29=data_naren.loc[(data_naren['Category']=='SPORTS')&(data_naren['Rating']),['App','Rating']].mean()
    x30=data_naren.loc[(data_naren['Category']=='TOOLS')&(data_naren['Rating']),['App','Rating']].mean()
    x31=data_naren.loc[(data_naren['Category']=='TRAVEL_AND_LOCAL')&(data_naren['Rating']),['App','Rating']].mean()
    x32=data_naren.loc[(data_naren['Category']=='VIDEO_PLAYERS')&(data_naren['Rating']),['App','Rating']].mean()
    x33=data_naren.loc[(data_naren['Category']=='WEATHER')&(data_naren['Rating']),['App','Rating']].mean()
    
    Label(screen28, text=x20, bg='yellow', fg='black').pack()
    Label(screen28, text=x21, bg='yellow', fg='black').pack()
    Label(screen28, text=x22, bg='yellow', fg='black').pack()
    Label(screen28, text=x23, bg='yellow', fg='black').pack()
    Label(screen28, text=x24, bg='yellow', fg='black').pack()
    Label(screen28, text=x25, bg='yellow', fg='black').pack()
    Label(screen28, text=x26, bg='yellow', fg='black').pack()
    Label(screen28, text=x27, bg='yellow', fg='black').pack()
    Label(screen28, text=x28, bg='yellow', fg='black').pack()
    Label(screen28, text=x29, bg='yellow', fg='black').pack()
    Label(screen28, text=x30, bg='yellow', fg='black').pack()
    Label(screen28, text=x31, bg='yellow', fg='black').pack()
    Label(screen28, text=x32, bg='yellow', fg='black').pack()
    Label(screen28, text=x33, bg='yellow', fg='black').pack()
    screen28.mainloop()
def function5():
    global screen8 
    screen8 = Toplevel(screen) 
    screen8.title("FUNCTION 5")
    adjustWindow(screen8)
      

    img = ImageTk.PhotoImage(Image.open('function5.jpg'))
    panel =Label(screen8, image = img)
    panel.pack(side = "bottom", fill = "both", expand = "yes")
    
    screen8.mainloop()

def function6():
    global screen9 
    screen9 = Toplevel(screen) 
    screen9.title("FUNCTION 5")
    adjustWindow(screen9)
      
    data_naren1=data_naren[(data_naren['Year']>=2015)&(data_naren['Year']<=2018)]
    plt.subplots(figsize=(6,6))
    plt.title("DOWNLOAD TREND CATEGORY WISE")
    sns.countplot(x="Category",hue='Year',data=data_naren1)
    plt.xticks(fontsize=6,rotation=78)
    plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function6.jpg',dpi=100)
  
    img = ImageTk.PhotoImage(Image.open('function6.jpg'))
    panel =Label(screen9, image = img)
    panel.pack(side="bottom", fill="both",expand="yes")
    
    screen9.mainloop()
    
def function7():
    global screen10 
    screen10 = Toplevel(screen) 
    screen10.title("FUNCTION 7")
    adjustWindow(screen10)
      
    plt.subplots(figsize=(6,6))
    plt.title("THE PERCENTAGE OF DOWNLOADS AS ANDROID VERSION")
    sns.countplot(data_naren['Android Ver'],label='percentage')
    plt.xticks(fontsize=6,rotation=78)
    plt.savefig('C:\\Users\\Admin\\Desktop\\internship gui\\function7.jpg',dpi=100)

    img = ImageTk.PhotoImage(Image.open('function7.jpg'))
    panel =Label(screen10, image = img)
    panel.pack(side="bottom", fill="both",expand="yes")
    
    screen10.mainloop()
    
def function8():
    global screen11 
    screen11 = Toplevel(screen) 
    screen11.title("FUNCTION 9")
    adjustWindow(screen11)
    
    img = ImageTk.PhotoImage(Image.open('function8.jpg'))
    panel =Label(screen11, image = img)
    panel.pack(side = "bottom", fill = "both", expand = "yes")
        
    screen11.mainloop()
    
def function9():
    global screen12 
    screen12 = Toplevel(screen) 
    screen12.title("FUNCTION 9")
    adjustWindow(screen12)
          
    l=data_naren.loc[(data_naren['Rating']==4.1)& (data_naren['Installs']=='100,000+'),['Installs','Rating','App']].count()
    
    Label(screen12, text="Apps with rating 4.1 & downloads 100,000+", bg='yellow', fg='black').pack()
    Label(screen12, text=l, bg='yellow', fg='black').pack()
    screen12.mainloop()

def function10():
    global screen13 
    screen13 = Toplevel(screen) 
    screen13.title("FUNCTION 9")
    adjustWindow(screen13)
    
    contentrating_list=data_harsh['Content Rating'].tolist()
    
    l=contentrating_list.count('Teen')
    l1=contentrating_list.count('Mature 17+')
    Label(screen13, text="Total number of teen rating apps downloaded", bg='yellow', fg='black').pack()
    Label(screen13, text=l, bg='yellow', fg='black').pack()
    Label(screen13, text="Total number of Mature 17+ rating apps downloaded", bg='yellow', fg='black').pack()
    Label(screen13, text=l1, bg='yellow', fg='black').pack()          
    img = ImageTk.PhotoImage(Image.open('function10.jpg'))
    panel =Label(screen13, image = img)
    panel.pack(side="bottom", fill="both",expand="yes")
    
    screen13.mainloop()
    
def function10():
    global screen13 
    screen13 = Toplevel(screen) 
    screen13.title("FUNCTION 9")
    adjustWindow(screen13)
    
    contentrating_list=data_harsh['Content Rating'].tolist()
    
    l=contentrating_list.count('Teen')
    l1=contentrating_list.count('Mature 17+')
    Label(screen13, text="Total number of teen rating apps downloaded", bg='yellow', fg='black').pack()
    Label(screen13, text=l, bg='yellow', fg='black').pack()
    Label(screen13, text="Total number of Mature 17+ rating apps downloaded", bg='yellow', fg='black').pack()
    Label(screen13, text=l1, bg='yellow', fg='black').pack()          
    img = ImageTk.PhotoImage(Image.open('function10.jpg'))
    panel =Label(screen13, image = img)
    panel.pack(side="bottom", fill="both",expand="yes")
    
    screen13.mainloop()
    
def function11():
    global screen14 
    screen14 = Toplevel(screen) 
    screen14.title("FUNCTION 11")
    adjustWindow(screen14)    
    
       
    img = ImageTk.PhotoImage(Image.open('function11.jpg'))
    panel =Label(screen14, image = img)
    panel.pack(side="bottom", fill="both",expand="yes")
    
    x10=data_naren.loc[(data_naren['Installs']=='500000')&(data_naren['quarter_year']=='2018Q3'),['Installs','App']]
    
    Label(screen14, text="APP WITH HIGHEST INSTALLS IN 2018Q3", bg='yellow', fg='black').pack()
    Label(screen14, text=x10, bg='yellow', fg='black').pack()
    
    screen14.mainloop()
    
def function12():
    global screen15 
    screen15 = Toplevel(screen) 
    screen15.title("FUNCTION 11")
    adjustWindow(screen15) 
    
    
    x1=data_priya.loc[(data_priya['App']=='10 Best Foods for You')&((data_priya['Sentiment']=='Positive')),['App']].count()

    x2=data_priya.loc[(data_priya['App']=='11st')&((data_priya['Sentiment']=='Positive')),['App']].count()

    x3=data_priya.loc[(data_priya['App']=='1800 Contacts - Lens Store')&((data_priya['Sentiment']=='Positive')),['App']].count()
    
    Label(screen15, text="10 Best Foods for You", bg='yellow', fg='black').pack()
    Label(screen15, text=x1, bg='yellow', fg='black').pack()
    Label(screen15, text="11st", bg='yellow', fg='black').pack()
    Label(screen15, text=x2, bg='yellow', fg='black').pack()
    Label(screen15, text="1800 Contacts - Lens Store", bg='yellow', fg='black').pack()
    Label(screen15, text=x3, bg='yellow', fg='black').pack()
    
    

    screen15.mainloop()
    
def function13():
    global screen17 
    screen17 = Toplevel(screen) 
    screen17.title("FUNCTION 13")
    adjustWindow(screen17) 
    
    x1=data_priya['Sentiment_Polarity'].corr(data_priya['Sentiment_Subjectivity'])
    Label(screen17, text="CORRELATION VALUE", bg='yellow', fg='black').pack()
    Label(screen17, text=x1, bg='yellow', fg='black').pack()
    screen17.mainloop()

def function14():
    global screen24,app_name,appname
    app_name=StringVar()
    screen24 = Toplevel(screen) 
    screen24.title("FUNCTION 14")
    adjustWindow(screen24) 
    
    Label(screen24, text="Enter name of the application", bg='yellow', fg='black').pack()
    Entry(screen24,textvar=app_name).pack()
        
    Button(screen24, text="Proceed", height="2", width="30", bg='#e79700', font=("Open Sans", 10, 'bold'), fg='white', command=function14sub).pack()
    
    screen24.mainloop()

def function14sub():
    global screen25
    screen25 = Toplevel(screen) 
    screen25.title("FUNCTION 14")
    adjustWindow(screen25)
    
    x1=data_priya.loc[(data_priya['App']==app_name.get())&(data_priya['Sentiment']=='Positive'),['App','Sentiment']].count()
    
    x2=data_priya.loc[(data_priya['App']==app_name.get())&(data_priya['Sentiment']=='Negative'),['App','Sentiment']].count()
    Label(screen25, text="Positive sentiments", bg='yellow', fg='black').pack()
    Label(screen25, text=x1, bg='yellow', fg='black').pack()
    Label(screen25, text="Negative sentiments", bg='yellow', fg='black').pack()
    Label(screen25, text=x2, bg='yellow', fg='black').pack()
    
    Label(screen25, text=x2, bg='yellow', fg='black').pack()
    
    screen25.mainloop()


def function15():
    global screen18 
    screen18 = Toplevel(screen) 
    screen18.title("FUNCTION 15")
    adjustWindow(screen18)

    
    sentiment_list=data_priya['Sentiment'].head(201).tolist()

    a1=sentiment_list.count('Positive')
    a2=sentiment_list.count('Negative')
    a3=sentiment_list.count('Neutral')
    a4=sentiment_list.count('nan')
    
    Label(screen18, text="Total number of Positive Sentiments", bg='yellow', fg='black').pack()
    Label(screen18, text=a1, bg='yellow', fg='black').pack()
    Label(screen18, text="Total number of Negative Sentiments", bg='yellow', fg='black').pack()
    Label(screen18, text=a2, bg='yellow', fg='black').pack()
    Label(screen18, text="Total number of Neutral Sentiments", bg='yellow', fg='black').pack()
    Label(screen18, text=a3, bg='yellow', fg='black').pack()
    Label(screen18, text="Total number of null sentimental data", bg='yellow', fg='black').pack()
    Label(screen18, text=a4, bg='yellow', fg='black').pack()
    
    y=int(str(len(data_priya.head(201))))    
    
    b1=(sentiment_list.count('Positive')/y)*100
    b2=(sentiment_list.count('Negative')/y)*100
    b3=(sentiment_list.count('Neutral')/y)*100
    
    
    Label(screen18, text="Percentage of Positive Sentiments", bg='yellow', fg='black').pack()
    Label(screen18, text=b1, bg='yellow', fg='black').pack()
    Label(screen18, text="Percentage of Negative Sentiments", bg='yellow', fg='black').pack()
    Label(screen18, text=b2, bg='yellow', fg='black').pack()
    Label(screen18, text="Percentage of Neutral Sentiments", bg='yellow', fg='black').pack()
    Label(screen18, text=b3, bg='yellow', fg='black').pack()

    
    screen18.mainloop()

def function16():
    global screen19 
    screen19 = Toplevel(screen) 
    screen19.title("FUNCTION 16")
    adjustWindow(screen19)

    b1=data_naren.groupby(['Year','Month','Installs'], )['Installs'].count()
    Label(screen19, text=b1, bg='yellow', fg='black').pack()
    #print(data_naren.groupby(['Year','Installs'], )['Installs'].count())
    screen19.mainloop()
    


def function17():
    global screen20 
    screen20 = Toplevel(screen) 
    screen20.title("FUNCTION 17")
    adjustWindow(screen20)
    
    x1=data_naren.groupby(['Installs','Size'], )['Installs'].count()

    Label(screen20, text=x1, bg='yellow', fg='black').pack()
    screen20.mainloop()
    
def function19():
    global screen22 
    screen22 = Toplevel(screen) 
    screen22.title("FUNCTION 19")
    adjustWindow(screen22)
    
    img = ImageTk.PhotoImage(Image.open('function19.jpg'))
    panel =Label(screen22, image = img)
    panel.pack(side = "bottom", fill = "both", expand = "yes")
    screen22.mainloop()

def function20():
    global screen23 
    screen23 = Toplevel(screen) 
    screen23.title("FUNCTION 19")
    adjustWindow(screen23)
    
    x1=data_naren.loc[(data_naren['Installs']=='5000000')&(data_naren['Type']=='Free')&(data_naren['Category']=='ART_AND_DESIGN')&(data_naren['Content Rating']=='Everyone'),['Installs','App','Content Rating']]
    Label(screen23, text=x1, bg='yellow', fg='black').pack()
    
    screen23.mainloop()
def insert(): 
    #rb=open_workbook("C:\\Datasets\\instenshipdata.xlsx",formatting_info='True')
    #wb = copy(rb)
    #f=open("C:\\Datasets\\instenshipdata.xlsx", "w+")
    current_row = sheet.max_row
    current_column = sheet.max_column
    
    sheet.cell(row=current_row, column=2).value=appname.get()  
    
       
       
        #sheet.cell(row=current_row+1, column=5).value = name_field.get()
        #sheet.cell(row=current_row+1, column=7).value = age_field.get()
        #sheet.cell(row=current_row+1, column=6).value = gender_field.get()
        #sheet.cell(row=current_row+1, column=12).value = height_field.get()
        #sheet.cell(row=current_row+1, column=11).value = weight_field.get()
        #sheet.cell(row=current_row, column=8).value = bloodgrp_field.get()
        #sheet.cell(row=current_row, column=15).value = bpsystolic_field.get()
        #sheet.cell(row=current_row, column=16).value = bpdystolic_field.get()
        #sheet.cell(row=current_row, column=10).value = haemoglobin_field.get()
        #sheet.cell(row=current_row, column=13).value = insulinBF_field.get()
        #sheet.cell(row=current_row, column=14).value = insulinAF_field.get()
        #sheet.cell(row=current_row,  column=2).value = diabetic_field.get()
    
        #sheet.cell(row=current_row+1, column=9).value = cost_field.get()
    wb.save("C:\\Datasets\\instenshipdata.xlsx")
        #name_field.focus_set()


main_screen() 
